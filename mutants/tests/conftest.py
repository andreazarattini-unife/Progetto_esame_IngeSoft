"""Shared test helpers."""

from pathlib import Path
from typing import Any

import openpyxl
import pytest


def _create_measurement_workbook(
    path: Path,
    *,
    coil_id: str = "COIL-001",
    date_time: str = "2026-07-09 10:30:00",
    nominal: float = 120.0,
    measurements: list[float] | None = None,
) -> Path:
    """Create a minimal valid measurement workbook for tests."""
    if measurements is None:
        measurements = [100.0, 110.0, 120.0]

    workbook = openpyxl.Workbook()
    values = workbook.active
    values.title = "Values"
    values["B2"] = date_time
    values["B4"] = nominal
    values["B5"] = coil_id

    lengthprofiles = workbook.create_sheet("LengthProfiles")
    lengthprofiles.cell(row=1, column=1).value = "Length"
    lengthprofiles.cell(row=1, column=2).value = "Measure"
    for index, measurement in enumerate(measurements, start=2):
        lengthprofiles.cell(row=index, column=1).value = index - 1
        lengthprofiles.cell(row=index, column=2).value = measurement

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


@pytest.fixture
def workbook_factory():
    """Return a helper that creates valid measurement workbooks."""
    return _create_measurement_workbook


def sheet_values(sheet: Any, row: int, columns: int) -> list[Any]:
    """Return values from one worksheet row."""
    return [sheet.cell(row=row, column=column).value for column in range(1, columns + 1)]
