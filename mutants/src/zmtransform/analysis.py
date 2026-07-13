"""Core analysis functions for zinc measurement workbooks."""

from __future__ import annotations

import statistics
from collections.abc import Iterable
from numbers import Real
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from zmtransform.models import AnalysisRow, MeasurementSummary, ProfileStatistics

RESULT_HEADERS = (
    "CoilID",
    "DateTime",
    "BS total length",
    "BS Nominal",
    "BS Avg",
    "BS St.Dev",
    "BS min",
    "BS max",
    "TS total length",
    "TS Nominal",
    "TS Avg",
    "TS St.Dev",
    "TS min",
    "TS max",
)


from mutmut.mutation.trampoline import MutantDict
from mutmut.mutation.trampoline import wrap_in_trampoline as _mutmut_mutated


class MeasurementFileError(ValueError):
    """Raised when a measurement workbook does not respect the expected format."""
mutants_x_find_input_files__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x_find_input_files__mutmut)
def find_input_files(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_orig(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_1(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = None
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_2(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(None)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_3(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted(None),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_4(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob(None)),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_5(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir * "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_6(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "XXBSXX").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_7(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "bs").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_8(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("XX*.xlsxXX")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_9(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.XLSX")),
        sorted((base_dir / "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_10(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted(None),
    )


def x_find_input_files__mutmut_11(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob(None)),
    )


def x_find_input_files__mutmut_12(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir * "TS").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_13(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "XXTSXX").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_14(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "ts").glob("*.xlsx")),
    )


def x_find_input_files__mutmut_15(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("XX*.xlsxXX")),
    )


def x_find_input_files__mutmut_16(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")),
        sorted((base_dir / "TS").glob("*.XLSX")),
    )

mutants_x_find_input_files__mutmut['_mutmut_orig'] = x_find_input_files__mutmut_orig # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_1'] = x_find_input_files__mutmut_1 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_2'] = x_find_input_files__mutmut_2 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_3'] = x_find_input_files__mutmut_3 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_4'] = x_find_input_files__mutmut_4 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_5'] = x_find_input_files__mutmut_5 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_6'] = x_find_input_files__mutmut_6 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_7'] = x_find_input_files__mutmut_7 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_8'] = x_find_input_files__mutmut_8 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_9'] = x_find_input_files__mutmut_9 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_10'] = x_find_input_files__mutmut_10 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_11'] = x_find_input_files__mutmut_11 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_12'] = x_find_input_files__mutmut_12 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_13'] = x_find_input_files__mutmut_13 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_14'] = x_find_input_files__mutmut_14 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_15'] = x_find_input_files__mutmut_15 # type: ignore # mutmut generated
mutants_x_find_input_files__mutmut['x_find_input_files__mutmut_16'] = x_find_input_files__mutmut_16 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x__numeric_values__mutmut)
def _numeric_values(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_orig(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_1(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = None
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_2(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is not None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_3(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            break
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_4(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) and not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_5(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_6(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                None
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_7(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(None)

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_8(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(None))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_9(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) <= 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_10(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 3:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def x__numeric_values__mutmut_11(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            None
        )

    return numeric_values


def x__numeric_values__mutmut_12(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "XXper calcolare la deviazione standardXX"
        )

    return numeric_values


def x__numeric_values__mutmut_13(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "PER CALCOLARE LA DEVIAZIONE STANDARD"
        )

    return numeric_values

mutants_x__numeric_values__mutmut['_mutmut_orig'] = x__numeric_values__mutmut_orig # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_1'] = x__numeric_values__mutmut_1 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_2'] = x__numeric_values__mutmut_2 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_3'] = x__numeric_values__mutmut_3 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_4'] = x__numeric_values__mutmut_4 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_5'] = x__numeric_values__mutmut_5 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_6'] = x__numeric_values__mutmut_6 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_7'] = x__numeric_values__mutmut_7 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_8'] = x__numeric_values__mutmut_8 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_9'] = x__numeric_values__mutmut_9 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_10'] = x__numeric_values__mutmut_10 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_11'] = x__numeric_values__mutmut_11 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_12'] = x__numeric_values__mutmut_12 # type: ignore # mutmut generated
mutants_x__numeric_values__mutmut['x__numeric_values__mutmut_13'] = x__numeric_values__mutmut_13 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x_calculate_statistics__mutmut)
def calculate_statistics(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_orig(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_1(
    values: Iterable[Any],
    source_file: str | Path = "XX<profile>XX",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_2(
    values: Iterable[Any],
    source_file: str | Path = "<PROFILE>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_3(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = None
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_4(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(None, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_5(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, None)
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_6(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_7(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, )
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_8(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(None))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_9(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=None,
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_10(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=None,
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_11(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=None,
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_12(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=None,
    )


def x_calculate_statistics__mutmut_13(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_14(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_15(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_16(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        )


def x_calculate_statistics__mutmut_17(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) * len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_18(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(None) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_19(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(None),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_20(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(None),
        maximum=max(profile_values),
    )


def x_calculate_statistics__mutmut_21(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(None),
    )

mutants_x_calculate_statistics__mutmut['_mutmut_orig'] = x_calculate_statistics__mutmut_orig # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_1'] = x_calculate_statistics__mutmut_1 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_2'] = x_calculate_statistics__mutmut_2 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_3'] = x_calculate_statistics__mutmut_3 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_4'] = x_calculate_statistics__mutmut_4 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_5'] = x_calculate_statistics__mutmut_5 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_6'] = x_calculate_statistics__mutmut_6 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_7'] = x_calculate_statistics__mutmut_7 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_8'] = x_calculate_statistics__mutmut_8 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_9'] = x_calculate_statistics__mutmut_9 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_10'] = x_calculate_statistics__mutmut_10 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_11'] = x_calculate_statistics__mutmut_11 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_12'] = x_calculate_statistics__mutmut_12 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_13'] = x_calculate_statistics__mutmut_13 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_14'] = x_calculate_statistics__mutmut_14 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_15'] = x_calculate_statistics__mutmut_15 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_16'] = x_calculate_statistics__mutmut_16 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_17'] = x_calculate_statistics__mutmut_17 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_18'] = x_calculate_statistics__mutmut_18 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_19'] = x_calculate_statistics__mutmut_19 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_20'] = x_calculate_statistics__mutmut_20 # type: ignore # mutmut generated
mutants_x_calculate_statistics__mutmut['x_calculate_statistics__mutmut_21'] = x_calculate_statistics__mutmut_21 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x_read_profile_values__mutmut)
def read_profile_values(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_orig(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_1(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = None
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_2(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = ""

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_3(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=None,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_4(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=None,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_5(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=None,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_6(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=None,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_7(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_8(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_9(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_10(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_11(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=3,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_12(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=2,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_13(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=3,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_14(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=False,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_15(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None or measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_16(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is not None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_17(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is not None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_18(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            break
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_19(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = None
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_20(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(None)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_21(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is not None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def x_read_profile_values__mutmut_22(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(None)

    return total_length, measurements

mutants_x_read_profile_values__mutmut['_mutmut_orig'] = x_read_profile_values__mutmut_orig # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_1'] = x_read_profile_values__mutmut_1 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_2'] = x_read_profile_values__mutmut_2 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_3'] = x_read_profile_values__mutmut_3 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_4'] = x_read_profile_values__mutmut_4 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_5'] = x_read_profile_values__mutmut_5 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_6'] = x_read_profile_values__mutmut_6 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_7'] = x_read_profile_values__mutmut_7 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_8'] = x_read_profile_values__mutmut_8 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_9'] = x_read_profile_values__mutmut_9 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_10'] = x_read_profile_values__mutmut_10 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_11'] = x_read_profile_values__mutmut_11 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_12'] = x_read_profile_values__mutmut_12 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_13'] = x_read_profile_values__mutmut_13 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_14'] = x_read_profile_values__mutmut_14 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_15'] = x_read_profile_values__mutmut_15 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_16'] = x_read_profile_values__mutmut_16 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_17'] = x_read_profile_values__mutmut_17 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_18'] = x_read_profile_values__mutmut_18 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_19'] = x_read_profile_values__mutmut_19 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_20'] = x_read_profile_values__mutmut_20 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_21'] = x_read_profile_values__mutmut_21 # type: ignore # mutmut generated
mutants_x_read_profile_values__mutmut['x_read_profile_values__mutmut_22'] = x_read_profile_values__mutmut_22 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x_read_measurement_file__mutmut)
def read_measurement_file(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_orig(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_1(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = None
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_2(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(None)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_3(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = None
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_4(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(None, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_5(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=None, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_6(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=None)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_7(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_8(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_9(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, )
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_10(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=False, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_11(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=False)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_12(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(None) from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_13(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "XXValuesXX" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_14(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_15(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "VALUES" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_16(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_17(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(None)
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_18(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "XXLengthProfilesXX" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_19(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "lengthprofiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_20(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LENGTHPROFILES" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_21(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_22(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(None)

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_23(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = None
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_24(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["XXValuesXX"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_25(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_26(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["VALUES"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_27(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = None
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_28(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["XXLengthProfilesXX"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_29(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["lengthprofiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_30(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LENGTHPROFILES"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_31(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = None
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_32(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(None, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_33(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, None)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_34(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_35(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, )
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_36(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = None

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_37(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(None, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_38(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, None)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_39(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_40(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, )

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_41(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = None
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_42(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["XXB5XX"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_43(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["b5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_44(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id not in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_45(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, "XXXX"):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_46(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(None)

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_47(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=None,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_48(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=None,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_49(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=None,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_50(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=None,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_51(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=None,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_52(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=None,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_53(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=None,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_54(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=None,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_55(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=None,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_56(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_57(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_58(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_59(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_60(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_61(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_62(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_63(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_64(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_65(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["XXB2XX"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_66(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["b2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_67(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["XXB4XX"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def x_read_measurement_file__mutmut_68(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["b4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()

mutants_x_read_measurement_file__mutmut['_mutmut_orig'] = x_read_measurement_file__mutmut_orig # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_1'] = x_read_measurement_file__mutmut_1 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_2'] = x_read_measurement_file__mutmut_2 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_3'] = x_read_measurement_file__mutmut_3 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_4'] = x_read_measurement_file__mutmut_4 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_5'] = x_read_measurement_file__mutmut_5 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_6'] = x_read_measurement_file__mutmut_6 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_7'] = x_read_measurement_file__mutmut_7 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_8'] = x_read_measurement_file__mutmut_8 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_9'] = x_read_measurement_file__mutmut_9 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_10'] = x_read_measurement_file__mutmut_10 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_11'] = x_read_measurement_file__mutmut_11 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_12'] = x_read_measurement_file__mutmut_12 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_13'] = x_read_measurement_file__mutmut_13 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_14'] = x_read_measurement_file__mutmut_14 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_15'] = x_read_measurement_file__mutmut_15 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_16'] = x_read_measurement_file__mutmut_16 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_17'] = x_read_measurement_file__mutmut_17 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_18'] = x_read_measurement_file__mutmut_18 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_19'] = x_read_measurement_file__mutmut_19 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_20'] = x_read_measurement_file__mutmut_20 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_21'] = x_read_measurement_file__mutmut_21 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_22'] = x_read_measurement_file__mutmut_22 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_23'] = x_read_measurement_file__mutmut_23 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_24'] = x_read_measurement_file__mutmut_24 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_25'] = x_read_measurement_file__mutmut_25 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_26'] = x_read_measurement_file__mutmut_26 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_27'] = x_read_measurement_file__mutmut_27 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_28'] = x_read_measurement_file__mutmut_28 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_29'] = x_read_measurement_file__mutmut_29 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_30'] = x_read_measurement_file__mutmut_30 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_31'] = x_read_measurement_file__mutmut_31 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_32'] = x_read_measurement_file__mutmut_32 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_33'] = x_read_measurement_file__mutmut_33 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_34'] = x_read_measurement_file__mutmut_34 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_35'] = x_read_measurement_file__mutmut_35 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_36'] = x_read_measurement_file__mutmut_36 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_37'] = x_read_measurement_file__mutmut_37 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_38'] = x_read_measurement_file__mutmut_38 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_39'] = x_read_measurement_file__mutmut_39 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_40'] = x_read_measurement_file__mutmut_40 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_41'] = x_read_measurement_file__mutmut_41 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_42'] = x_read_measurement_file__mutmut_42 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_43'] = x_read_measurement_file__mutmut_43 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_44'] = x_read_measurement_file__mutmut_44 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_45'] = x_read_measurement_file__mutmut_45 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_46'] = x_read_measurement_file__mutmut_46 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_47'] = x_read_measurement_file__mutmut_47 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_48'] = x_read_measurement_file__mutmut_48 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_49'] = x_read_measurement_file__mutmut_49 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_50'] = x_read_measurement_file__mutmut_50 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_51'] = x_read_measurement_file__mutmut_51 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_52'] = x_read_measurement_file__mutmut_52 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_53'] = x_read_measurement_file__mutmut_53 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_54'] = x_read_measurement_file__mutmut_54 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_55'] = x_read_measurement_file__mutmut_55 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_56'] = x_read_measurement_file__mutmut_56 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_57'] = x_read_measurement_file__mutmut_57 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_58'] = x_read_measurement_file__mutmut_58 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_59'] = x_read_measurement_file__mutmut_59 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_60'] = x_read_measurement_file__mutmut_60 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_61'] = x_read_measurement_file__mutmut_61 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_62'] = x_read_measurement_file__mutmut_62 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_63'] = x_read_measurement_file__mutmut_63 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_64'] = x_read_measurement_file__mutmut_64 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_65'] = x_read_measurement_file__mutmut_65 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_66'] = x_read_measurement_file__mutmut_66 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_67'] = x_read_measurement_file__mutmut_67 # type: ignore # mutmut generated
mutants_x_read_measurement_file__mutmut['x_read_measurement_file__mutmut_68'] = x_read_measurement_file__mutmut_68 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x_analyze_measurements__mutmut)
def analyze_measurements(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_orig(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_1(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = None
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_2(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = None
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_3(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = None

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_4(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = None
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_5(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(None)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_6(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = None
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_7(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError:
            errors.append(None)

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_8(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = None
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_9(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(None)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_10(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                None
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_11(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=None,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_12(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=None,
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_13(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_14(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_15(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(None),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def x_analyze_measurements__mutmut_16(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError:
            errors.append(None)

    return result_rows, errors

mutants_x_analyze_measurements__mutmut['_mutmut_orig'] = x_analyze_measurements__mutmut_orig # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_1'] = x_analyze_measurements__mutmut_1 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_2'] = x_analyze_measurements__mutmut_2 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_3'] = x_analyze_measurements__mutmut_3 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_4'] = x_analyze_measurements__mutmut_4 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_5'] = x_analyze_measurements__mutmut_5 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_6'] = x_analyze_measurements__mutmut_6 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_7'] = x_analyze_measurements__mutmut_7 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_8'] = x_analyze_measurements__mutmut_8 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_9'] = x_analyze_measurements__mutmut_9 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_10'] = x_analyze_measurements__mutmut_10 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_11'] = x_analyze_measurements__mutmut_11 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_12'] = x_analyze_measurements__mutmut_12 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_13'] = x_analyze_measurements__mutmut_13 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_14'] = x_analyze_measurements__mutmut_14 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_15'] = x_analyze_measurements__mutmut_15 # type: ignore # mutmut generated
mutants_x_analyze_measurements__mutmut['x_analyze_measurements__mutmut_16'] = x_analyze_measurements__mutmut_16 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x_create_result_workbook__mutmut)
def create_result_workbook(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_orig(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_1(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = None
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_2(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = None
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_3(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = None

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_4(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "XXDatiXX"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_5(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_6(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "DATI"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_7(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(None, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_8(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=None):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_9(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_10(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, ):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_11(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=2):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_12(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = None

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_13(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=None, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_14(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=None).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_15(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_16(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, ).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_17(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=2).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_18(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(None, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_19(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=None):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_20(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_21(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, ):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_22(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=3):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_23(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = None
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_24(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = None

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_25(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = None
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_26(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=None, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_27(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=None).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_28(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_29(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, ).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_30(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=2, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_31(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = None
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_32(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=None, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_33(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=None).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_34(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_35(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, ).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_36(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=3, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_37(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = None
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_38(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=None, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_39(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=None).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_40(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_41(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, ).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_42(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=4, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_43(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = None
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_44(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=None, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_45(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=None).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_46(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_47(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, ).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_48(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=5, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_49(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = None
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_50(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=None, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_51(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=None).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_52(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_53(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, ).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_54(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=6, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_55(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = None
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_56(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=None, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_57(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=None).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_58(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_59(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, ).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_60(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=7, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_61(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = None
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_62(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=None, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_63(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=None).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_64(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_65(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, ).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_66(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=8, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_67(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = None

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_68(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=None, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_69(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=None).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_70(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_71(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, ).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_72(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=9, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_73(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_74(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = None
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_75(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=None, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_76(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=None).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_77(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_78(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, ).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_79(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=10, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_80(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = None
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_81(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=None, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_82(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=None).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_83(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_84(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, ).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_85(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=11, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_86(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = None
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_87(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=None, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_88(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=None).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_89(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_90(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, ).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_91(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=12, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_92(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = None
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_93(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=None, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_94(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=None).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_95(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_96(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, ).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_97(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=13, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_98(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = None
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_99(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=None, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_100(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=None).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_101(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_102(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, ).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_103(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=14, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_104(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = None

    return result_excel


def x_create_result_workbook__mutmut_105(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=None, row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_106(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=None).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_107(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(row=row).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_108(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, ).value = ts_measurement.maximum

    return result_excel


def x_create_result_workbook__mutmut_109(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=15, row=row).value = ts_measurement.maximum

    return result_excel

mutants_x_create_result_workbook__mutmut['_mutmut_orig'] = x_create_result_workbook__mutmut_orig # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_1'] = x_create_result_workbook__mutmut_1 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_2'] = x_create_result_workbook__mutmut_2 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_3'] = x_create_result_workbook__mutmut_3 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_4'] = x_create_result_workbook__mutmut_4 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_5'] = x_create_result_workbook__mutmut_5 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_6'] = x_create_result_workbook__mutmut_6 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_7'] = x_create_result_workbook__mutmut_7 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_8'] = x_create_result_workbook__mutmut_8 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_9'] = x_create_result_workbook__mutmut_9 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_10'] = x_create_result_workbook__mutmut_10 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_11'] = x_create_result_workbook__mutmut_11 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_12'] = x_create_result_workbook__mutmut_12 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_13'] = x_create_result_workbook__mutmut_13 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_14'] = x_create_result_workbook__mutmut_14 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_15'] = x_create_result_workbook__mutmut_15 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_16'] = x_create_result_workbook__mutmut_16 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_17'] = x_create_result_workbook__mutmut_17 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_18'] = x_create_result_workbook__mutmut_18 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_19'] = x_create_result_workbook__mutmut_19 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_20'] = x_create_result_workbook__mutmut_20 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_21'] = x_create_result_workbook__mutmut_21 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_22'] = x_create_result_workbook__mutmut_22 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_23'] = x_create_result_workbook__mutmut_23 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_24'] = x_create_result_workbook__mutmut_24 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_25'] = x_create_result_workbook__mutmut_25 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_26'] = x_create_result_workbook__mutmut_26 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_27'] = x_create_result_workbook__mutmut_27 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_28'] = x_create_result_workbook__mutmut_28 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_29'] = x_create_result_workbook__mutmut_29 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_30'] = x_create_result_workbook__mutmut_30 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_31'] = x_create_result_workbook__mutmut_31 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_32'] = x_create_result_workbook__mutmut_32 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_33'] = x_create_result_workbook__mutmut_33 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_34'] = x_create_result_workbook__mutmut_34 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_35'] = x_create_result_workbook__mutmut_35 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_36'] = x_create_result_workbook__mutmut_36 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_37'] = x_create_result_workbook__mutmut_37 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_38'] = x_create_result_workbook__mutmut_38 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_39'] = x_create_result_workbook__mutmut_39 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_40'] = x_create_result_workbook__mutmut_40 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_41'] = x_create_result_workbook__mutmut_41 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_42'] = x_create_result_workbook__mutmut_42 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_43'] = x_create_result_workbook__mutmut_43 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_44'] = x_create_result_workbook__mutmut_44 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_45'] = x_create_result_workbook__mutmut_45 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_46'] = x_create_result_workbook__mutmut_46 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_47'] = x_create_result_workbook__mutmut_47 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_48'] = x_create_result_workbook__mutmut_48 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_49'] = x_create_result_workbook__mutmut_49 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_50'] = x_create_result_workbook__mutmut_50 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_51'] = x_create_result_workbook__mutmut_51 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_52'] = x_create_result_workbook__mutmut_52 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_53'] = x_create_result_workbook__mutmut_53 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_54'] = x_create_result_workbook__mutmut_54 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_55'] = x_create_result_workbook__mutmut_55 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_56'] = x_create_result_workbook__mutmut_56 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_57'] = x_create_result_workbook__mutmut_57 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_58'] = x_create_result_workbook__mutmut_58 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_59'] = x_create_result_workbook__mutmut_59 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_60'] = x_create_result_workbook__mutmut_60 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_61'] = x_create_result_workbook__mutmut_61 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_62'] = x_create_result_workbook__mutmut_62 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_63'] = x_create_result_workbook__mutmut_63 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_64'] = x_create_result_workbook__mutmut_64 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_65'] = x_create_result_workbook__mutmut_65 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_66'] = x_create_result_workbook__mutmut_66 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_67'] = x_create_result_workbook__mutmut_67 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_68'] = x_create_result_workbook__mutmut_68 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_69'] = x_create_result_workbook__mutmut_69 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_70'] = x_create_result_workbook__mutmut_70 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_71'] = x_create_result_workbook__mutmut_71 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_72'] = x_create_result_workbook__mutmut_72 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_73'] = x_create_result_workbook__mutmut_73 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_74'] = x_create_result_workbook__mutmut_74 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_75'] = x_create_result_workbook__mutmut_75 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_76'] = x_create_result_workbook__mutmut_76 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_77'] = x_create_result_workbook__mutmut_77 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_78'] = x_create_result_workbook__mutmut_78 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_79'] = x_create_result_workbook__mutmut_79 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_80'] = x_create_result_workbook__mutmut_80 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_81'] = x_create_result_workbook__mutmut_81 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_82'] = x_create_result_workbook__mutmut_82 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_83'] = x_create_result_workbook__mutmut_83 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_84'] = x_create_result_workbook__mutmut_84 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_85'] = x_create_result_workbook__mutmut_85 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_86'] = x_create_result_workbook__mutmut_86 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_87'] = x_create_result_workbook__mutmut_87 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_88'] = x_create_result_workbook__mutmut_88 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_89'] = x_create_result_workbook__mutmut_89 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_90'] = x_create_result_workbook__mutmut_90 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_91'] = x_create_result_workbook__mutmut_91 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_92'] = x_create_result_workbook__mutmut_92 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_93'] = x_create_result_workbook__mutmut_93 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_94'] = x_create_result_workbook__mutmut_94 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_95'] = x_create_result_workbook__mutmut_95 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_96'] = x_create_result_workbook__mutmut_96 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_97'] = x_create_result_workbook__mutmut_97 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_98'] = x_create_result_workbook__mutmut_98 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_99'] = x_create_result_workbook__mutmut_99 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_100'] = x_create_result_workbook__mutmut_100 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_101'] = x_create_result_workbook__mutmut_101 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_102'] = x_create_result_workbook__mutmut_102 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_103'] = x_create_result_workbook__mutmut_103 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_104'] = x_create_result_workbook__mutmut_104 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_105'] = x_create_result_workbook__mutmut_105 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_106'] = x_create_result_workbook__mutmut_106 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_107'] = x_create_result_workbook__mutmut_107 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_108'] = x_create_result_workbook__mutmut_108 # type: ignore # mutmut generated
mutants_x_create_result_workbook__mutmut['x_create_result_workbook__mutmut_109'] = x_create_result_workbook__mutmut_109 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut: MutantDict = {}  # type: ignore


@_mutmut_mutated(mutants_x_analyze_directory_to_workbook__mutmut)
def analyze_directory_to_workbook(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(bs_files, ts_files)
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_orig(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(bs_files, ts_files)
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_1(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = None
    rows, errors = analyze_measurements(bs_files, ts_files)
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_2(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(None)
    rows, errors = analyze_measurements(bs_files, ts_files)
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_3(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = None
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_4(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(None, ts_files)
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_5(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(bs_files, None)
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_6(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(ts_files)
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_7(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(bs_files, )
    return create_result_workbook(rows), errors


def x_analyze_directory_to_workbook__mutmut_8(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(bs_files, ts_files)
    return create_result_workbook(None), errors

mutants_x_analyze_directory_to_workbook__mutmut['_mutmut_orig'] = x_analyze_directory_to_workbook__mutmut_orig # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_1'] = x_analyze_directory_to_workbook__mutmut_1 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_2'] = x_analyze_directory_to_workbook__mutmut_2 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_3'] = x_analyze_directory_to_workbook__mutmut_3 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_4'] = x_analyze_directory_to_workbook__mutmut_4 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_5'] = x_analyze_directory_to_workbook__mutmut_5 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_6'] = x_analyze_directory_to_workbook__mutmut_6 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_7'] = x_analyze_directory_to_workbook__mutmut_7 # type: ignore # mutmut generated
mutants_x_analyze_directory_to_workbook__mutmut['x_analyze_directory_to_workbook__mutmut_8'] = x_analyze_directory_to_workbook__mutmut_8 # type: ignore # mutmut generated
