"""Core analysis functions for zinc measurement workbooks."""

from __future__ import annotations

import statistics
from collections.abc import Iterable
from numbers import Real
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from zmtransform.models import AnalysisRow, MeasurementSummary, ProfileStatistics

RESULT_HEADERS = (
    "CoilID",
    "DateTime",
    "BS total length",
    "BS Nominal",
    "BS Avg",
    "BS St.Dev",
    "BS min",
    "BS max",
    "TS total length",
    "TS Nominal",
    "TS Avg",
    "TS St.Dev",
    "TS min",
    "TS max",
)


class MeasurementFileError(ValueError):
    """Raised when a measurement workbook does not respect the expected format."""


def find_input_files(input_dir: str | Path) -> tuple[list[Path], list[Path]]:
    """Mette in ordine gli excel di BS e TS contenuti nella cartella passata come parametro"""
    base_dir = Path(input_dir)
    return (
        sorted((base_dir / "BS").glob("*.xlsx")), #qui vado a ordinre i file excel di BS
        sorted((base_dir / "TS").glob("*.xlsx")), #qui vado a fare la stessa cosa per TS
    )


def _numeric_values(values: Iterable[Any], source_file: Path) -> list[float]:
    numeric_values: list[float] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, bool) or not isinstance(value, Real):
            raise MeasurementFileError(
                f"{source_file}: valore non numerico in LengthProfiles colonna 2: {value!r}"
            )
        numeric_values.append(float(value))

    if len(numeric_values) < 2:
        raise MeasurementFileError(
            f"{source_file}: servono almeno due misure numeriche "
            "per calcolare la deviazione standard"
        )

    return numeric_values


def calculate_statistics(
    values: Iterable[Any],
    source_file: str | Path = "<profile>",
) -> ProfileStatistics:
    """Calculate descriptive statistics for a measurement profile."""
    profile_values = _numeric_values(values, Path(source_file))
    return ProfileStatistics(
        average=sum(profile_values) / len(profile_values),
        standard_deviation=statistics.stdev(profile_values),
        minimum=min(profile_values),
        maximum=max(profile_values),
    )


def read_profile_values(lengthprofiles_sheet: Any, source_file: Path) -> tuple[Any, list[Any]]:
    """Read total length and measurement values from a LengthProfiles worksheet."""
    measurements: list[Any] = []
    total_length = None

    for length, measurement in lengthprofiles_sheet.iter_rows(
        min_row=2,
        min_col=1,
        max_col=2,
        values_only=True,
    ):
        if length is None and measurement is None:
            continue
        total_length = length
        measurements.append(measurement)

    if total_length is None:
        raise MeasurementFileError(f"{source_file}: nessuna riga dati in LengthProfiles")

    return total_length, measurements


def read_measurement_file(file_path: str | Path) -> MeasurementSummary:
    """Read one BS/TS Excel file and return its measurement summary."""
    source_file = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    except (OSError, InvalidFileException, BadZipFile) as exc:
        raise MeasurementFileError(f"{source_file}: impossibile aprire il file Excel") from exc

    try:
        if "Values" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'Values' mancante")
        if "LengthProfiles" not in workbook.sheetnames:
            raise MeasurementFileError(f"{source_file}: foglio 'LengthProfiles' mancante")

        values_sheet = workbook["Values"]
        lengthprofiles_sheet = workbook["LengthProfiles"]
        total_length, profile_values = read_profile_values(lengthprofiles_sheet, source_file)
        statistics_summary = calculate_statistics(profile_values, source_file)

        coil_id = values_sheet["B5"].value
        if coil_id in (None, ""):
            raise MeasurementFileError(f"{source_file}: CoilID mancante in Values!B5")

        return MeasurementSummary(
            coil_id=coil_id,
            date_time=values_sheet["B2"].value,
            nominal=values_sheet["B4"].value,
            total_length=total_length,
            average=statistics_summary.average,
            standard_deviation=statistics_summary.standard_deviation,
            minimum=statistics_summary.minimum,
            maximum=statistics_summary.maximum,
            source_file=source_file,
        )
    finally:
        workbook.close()


def analyze_measurements(
    data_bs: Iterable[str | Path],
    data_ts: Iterable[str | Path],
) -> tuple[list[AnalysisRow], list[str]]:
    """Analyze BS files and match TS data by CoilID."""
    errors: list[str] = []
    result_rows: list[AnalysisRow] = []
    ts_measurements: dict[Any, MeasurementSummary] = {}

    for file_path in data_ts:
        try:
            ts_measurement = read_measurement_file(file_path)
            ts_measurements[ts_measurement.coil_id] = ts_measurement
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK FOR TS - {exc}")

    for file_path in data_bs:
        try:
            bs_measurement = read_measurement_file(file_path)
            result_rows.append(
                AnalysisRow(
                    bs=bs_measurement,
                    ts=ts_measurements.get(bs_measurement.coil_id),
                )
            )
        except MeasurementFileError as exc:
            errors.append(f"{file_path}: NOT OK - {exc}")

    return result_rows, errors


def create_result_workbook(measurement_rows: Iterable[AnalysisRow]) -> openpyxl.Workbook:
    """Create the final Excel workbook from analyzed measurement rows."""
    result_excel = openpyxl.Workbook()
    result_sheet = result_excel.active
    result_sheet.title = "Dati"

    for column, header in enumerate(RESULT_HEADERS, start=1):
        result_sheet.cell(column=column, row=1).value = header

    for row, measurement_row in enumerate(measurement_rows, start=2):
        bs_measurement = measurement_row.bs
        ts_measurement = measurement_row.ts

        result_sheet.cell(column=1, row=row).value = bs_measurement.coil_id
        result_sheet.cell(column=2, row=row).value = bs_measurement.date_time
        result_sheet.cell(column=3, row=row).value = bs_measurement.total_length
        result_sheet.cell(column=4, row=row).value = bs_measurement.nominal
        result_sheet.cell(column=5, row=row).value = bs_measurement.average
        result_sheet.cell(column=6, row=row).value = bs_measurement.standard_deviation
        result_sheet.cell(column=7, row=row).value = bs_measurement.minimum
        result_sheet.cell(column=8, row=row).value = bs_measurement.maximum

        if ts_measurement is not None:
            result_sheet.cell(column=9, row=row).value = ts_measurement.total_length
            result_sheet.cell(column=10, row=row).value = ts_measurement.nominal
            result_sheet.cell(column=11, row=row).value = ts_measurement.average
            result_sheet.cell(column=12, row=row).value = ts_measurement.standard_deviation
            result_sheet.cell(column=13, row=row).value = ts_measurement.minimum
            result_sheet.cell(column=14, row=row).value = ts_measurement.maximum

    return result_excel


def analyze_directory_to_workbook(input_dir: str | Path) -> tuple[openpyxl.Workbook, list[str]]:
    """Analyze a directory with BS/TS subfolders and return the output workbook."""
    bs_files, ts_files = find_input_files(input_dir)
    rows, errors = analyze_measurements(bs_files, ts_files)
    return create_result_workbook(rows), errors
